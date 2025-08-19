from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
import os
from functools import wraps
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///posyandu.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# Models
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(120), nullable=False)
    full_name = db.Column(db.String(100), nullable=False)
    role = db.Column(db.String(20), default='user')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

class PosyanduData(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    dusun = db.Column(db.String(100), nullable=False)
    nama = db.Column(db.String(100), nullable=False)
    kategori = db.Column(db.String(20), nullable=False)  # Only 'Balita' and 'Lansia'
    jenis_kelamin = db.Column(db.String(10), nullable=False)
    tanggal_lahir = db.Column(db.Date, nullable=False)
    umur = db.Column(db.Integer, nullable=False)
    alamat = db.Column(db.Text, nullable=False)
    nomor_ktp = db.Column(db.String(16))
    nomor_bpjs = db.Column(db.String(20))
    berat_badan = db.Column(db.Float, nullable=False)
    tinggi_badan = db.Column(db.Float, nullable=False)
    imt = db.Column(db.Float)
    lingkar_perut = db.Column(db.Float)
    lingkar_kepala = db.Column(db.Float)
    lila = db.Column(db.Float)
    tekanan_darah = db.Column(db.String(20))
    mental_dan_emosional = db.Column(db.Text)
    keterangan = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)

    def calculate_imt(self):
        if self.berat_badan and self.tinggi_badan:
            tinggi_meter = self.tinggi_badan / 100
            self.imt = round(self.berat_badan / (tinggi_meter * tinggi_meter), 1)

    def get_imt_status(self):
        if not self.imt:
            return ''
        if self.imt < 18.5:
            return 'Underweight'
        elif self.imt < 25:
            return 'Normal'
        elif self.imt < 30:
            return 'Overweight'
        else:
            return 'Obese'

    def get_imt_color(self):
        if not self.imt:
            return 'secondary'
        if self.imt < 18.5:
            return 'info'
        elif self.imt < 25:
            return 'success'
        elif self.imt < 30:
            return 'warning'
        else:
            return 'danger'

# Login required decorator
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Silakan login terlebih dahulu.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Initialize database function
def init_db():
    """Initialize database and create default admin user"""
    with app.app_context():
        db.create_all()
        
        # Create default admin user if not exists
        if not User.query.filter_by(username='admin').first():
            admin = User(
                username='admin',
                email='admin@posyandu.com',
                full_name='Administrator',
                role='admin'
            )
            admin.set_password('admin123')
            db.session.add(admin)
            db.session.commit()
            print("✅ Default admin user created!")
            print("Username: admin")
            print("Password: admin123")

# Routes
@app.route('/')
def index():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return redirect(url_for('dashboard'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            session['user_id'] = user.id
            session['username'] = user.username
            session['full_name'] = user.full_name
            flash(f'Selamat datang, {user.full_name}!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Username atau password salah.', 'danger')
    
    return render_template('auth/login.html')

@app.route('/logout')
def logout():
    user_name = session.get('full_name', 'User')
    session.clear()
    flash(f'Sampai jumpa, {user_name}! Anda telah logout.', 'info')
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    total_data = PosyanduData.query.count()
    recent_data = PosyanduData.query.order_by(PosyanduData.created_at.desc()).limit(5).all()
    
    # Statistics by category (only Balita and Lansia)
    stats = {
        'Balita': PosyanduData.query.filter_by(kategori='Balita').count(),
        'Lansia': PosyanduData.query.filter_by(kategori='Lansia').count(),
    }
    
    return render_template('dashboard.html', 
                         total_data=total_data, 
                         recent_data=recent_data, 
                         stats=stats)

@app.route('/form')
@login_required
def form():
    return render_template('form.html')

@app.route('/submit', methods=['POST'])
@login_required
def submit():
    try:
        # Validate kategori
        kategori = request.form['kategori']
        if kategori not in ['Balita', 'Lansia']:
            flash('Kategori tidak valid. Hanya Balita dan Lansia yang diperbolehkan.', 'danger')
            return redirect(url_for('form'))
        
        data = PosyanduData(
            dusun=request.form['dusun'],
            nama=request.form['nama'],
            kategori=kategori,
            jenis_kelamin=request.form['jenis_kelamin'],
            tanggal_lahir=datetime.strptime(request.form['tanggal_lahir'], '%Y-%m-%d').date(),
            umur=int(request.form['umur']),
            alamat=request.form['alamat'],
            nomor_ktp=request.form.get('nomor_ktp', ''),
            nomor_bpjs=request.form.get('nomor_bpjs', ''),
            berat_badan=float(request.form['berat_badan']),
            tinggi_badan=float(request.form['tinggi_badan']),
            lingkar_perut=float(request.form['lingkar_perut']) if request.form.get('lingkar_perut') else None,
            lingkar_kepala=float(request.form['lingkar_kepala']) if request.form.get('lingkar_kepala') else None,
            lila=float(request.form['lila']) if request.form.get('lila') else None,
            tekanan_darah=request.form.get('tekanan_darah', ''),
            mental_dan_emosional=request.form.get('mental_dan_emosional', ''),
            keterangan=request.form.get('keterangan', ''),
            created_by=session['user_id']
        )
        
        data.calculate_imt()
        db.session.add(data)
        db.session.commit()
        
        flash('Data berhasil disimpan!', 'success')
        return redirect(url_for('result', id=data.id))
    
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('form'))

@app.route('/result/<int:id>')
@login_required
def result(id):
    data = PosyanduData.query.get_or_404(id)
    return render_template('result.html', data=data)

@app.route('/riwayat')
@login_required
def riwayat():
    search = request.args.get('search', '')
    kategori = request.args.get('kategori', '')
    page = request.args.get('page', 1, type=int)
    
    query = PosyanduData.query
    
    if search:
        query = query.filter(
            db.or_(
                PosyanduData.nama.contains(search),
                PosyanduData.dusun.contains(search),
                PosyanduData.alamat.contains(search)
            )
        )
    
    if kategori and kategori in ['Balita', 'Lansia']:
        query = query.filter_by(kategori=kategori)
    
    data = query.order_by(PosyanduData.created_at.desc()).paginate(
        page=page, per_page=10, error_out=False
    )
    
    return render_template('riwayat.html', data=data, search=search, kategori=kategori)

@app.route('/edit/<int:id>')
@login_required
def edit(id):
    data = PosyanduData.query.get_or_404(id)
    return render_template('edit.html', data=data)

@app.route('/update/<int:id>', methods=['POST'])
@login_required
def update(id):
    data = PosyanduData.query.get_or_404(id)
    
    try:
        # Validate kategori
        kategori = request.form['kategori']
        if kategori not in ['Balita', 'Lansia']:
            flash('Kategori tidak valid. Hanya Balita dan Lansia yang diperbolehkan.', 'danger')
            return redirect(url_for('edit', id=id))
        
        data.dusun = request.form['dusun']
        data.nama = request.form['nama']
        data.kategori = kategori
        data.jenis_kelamin = request.form['jenis_kelamin']
        data.tanggal_lahir = datetime.strptime(request.form['tanggal_lahir'], '%Y-%m-%d').date()
        data.umur = int(request.form['umur'])
        data.alamat = request.form['alamat']
        data.nomor_ktp = request.form.get('nomor_ktp', '')
        data.nomor_bpjs = request.form.get('nomor_bpjs', '')
        data.berat_badan = float(request.form['berat_badan'])
        data.tinggi_badan = float(request.form['tinggi_badan'])
        data.lingkar_perut = float(request.form['lingkar_perut']) if request.form.get('lingkar_perut') else None
        data.lingkar_kepala = float(request.form['lingkar_kepala']) if request.form.get('lingkar_kepala') else None
        data.lila = float(request.form['lila']) if request.form.get('lila') else None
        data.tekanan_darah = request.form.get('tekanan_darah', '')
        data.mental_dan_emosional = request.form.get('mental_dan_emosional', '')
        data.keterangan = request.form.get('keterangan', '')
        data.updated_at = datetime.utcnow()
        
        data.calculate_imt()
        db.session.commit()
        
        flash('Data berhasil diperbarui!', 'success')
        return redirect(url_for('riwayat'))
    
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('edit', id=id))

@app.route('/delete/<int:id>', methods=['POST'])
@login_required
def delete(id):
    data = PosyanduData.query.get_or_404(id)
    nama = data.nama
    
    db.session.delete(data)
    db.session.commit()
    
    flash(f'Data {nama} berhasil dihapus!', 'success')
    return redirect(url_for('riwayat'))

@app.route('/export/excel')
@login_required
def export_excel():
    try:
        # Get filter parameters
        search = request.args.get('search', '')
        kategori = request.args.get('kategori', '')
        
        # Build query with same filters as riwayat page
        query = PosyanduData.query
        
        if search:
            query = query.filter(
                db.or_(
                    PosyanduData.nama.contains(search),
                    PosyanduData.dusun.contains(search),
                    PosyanduData.alamat.contains(search)
                )
            )
        
        if kategori and kategori in ['Balita', 'Lansia']:
            query = query.filter_by(kategori=kategori)
        
        data = query.order_by(PosyanduData.created_at.desc()).all()
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Posyandu"
        
        # Define headers
        headers = [
            'No', 'Dusun', 'Nama Lengkap', 'Kategori', 'Jenis Kelamin', 
            'Tanggal Lahir', 'Umur', 'Alamat', 'Nomor KTP', 'Nomor BPJS',
            'Berat Badan (kg)', 'Tinggi Badan (cm)', 'IMT', 'Status IMT',
            'Lingkar Perut (cm)', 'Lingkar Kepala (cm)', 'Lila (cm)',
            'Tekanan Darah (mmHg)', 'Mental & Emosional',
            'Keterangan', 'Tanggal Input'
        ]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data rows
        for row_idx, item in enumerate(data, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1)  # No
            ws.cell(row=row_idx, column=2, value=item.dusun)
            ws.cell(row=row_idx, column=3, value=item.nama)
            ws.cell(row=row_idx, column=4, value=item.kategori)
            ws.cell(row=row_idx, column=5, value=item.jenis_kelamin)
            ws.cell(row=row_idx, column=6, value=item.tanggal_lahir.strftime('%d/%m/%Y'))
            ws.cell(row=row_idx, column=7, value=item.umur)
            ws.cell(row=row_idx, column=8, value=item.alamat)
            ws.cell(row=row_idx, column=9, value=item.nomor_ktp or '')
            ws.cell(row=row_idx, column=10, value=item.nomor_bpjs or '')
            ws.cell(row=row_idx, column=11, value=item.berat_badan)
            ws.cell(row=row_idx, column=12, value=item.tinggi_badan)
            ws.cell(row=row_idx, column=13, value=item.imt or '')
            ws.cell(row=row_idx, column=14, value=item.get_imt_status())
            ws.cell(row=row_idx, column=15, value=item.lingkar_perut or '')
            ws.cell(row=row_idx, column=16, value=item.lingkar_kepala or '')
            ws.cell(row=row_idx, column=17, value=item.lila or '')
            ws.cell(row=row_idx, column=18, value=item.tekanan_darah or '')
            ws.cell(row=row_idx, column=19, value=item.mental_dan_emosional or '')
            ws.cell(row=row_idx, column=20, value=item.keterangan or '')
            ws.cell(row=row_idx, column=21, value=item.created_at.strftime('%d/%m/%Y %H:%M'))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Max width 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'data_posyandu_{timestamp}.xlsx'
        
        flash(f'Data berhasil diekspor! Total: {len(data)} record', 'success')
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'Error saat export Excel: {str(e)}', 'danger')
        return redirect(url_for('riwayat'))

@app.route('/api/calculate-imt', methods=['POST'])
@login_required
def calculate_imt_api():
    try:
        berat_badan = float(request.json.get('berat_badan', 0))
        tinggi_badan = float(request.json.get('tinggi_badan', 0))
        
        if berat_badan > 0 and tinggi_badan > 0:
            tinggi_meter = tinggi_badan / 100
            imt = round(berat_badan / (tinggi_meter * tinggi_meter), 1)
            return jsonify({'imt': imt})
        
        return jsonify({'imt': 0})
    except:
        return jsonify({'imt': 0})

if __name__ == '__main__':
    # Initialize database when app starts
    init_db()
    
    print("🚀 Starting Posyandu System...")
    print("📍 Open browser: http://localhost:5000")
    print("👤 Login: admin / admin123")
    print("📊 Categories: Balita & Lansia only")
    print("📥 Excel export available")
    
    app.run(debug=True, host='0.0.0.0', port=5000)
